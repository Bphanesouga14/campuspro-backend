#Importer le fichier Excel complet


#  RÔLE : Use case d'import du fichier Excel vers la base.
#
#  SCÉNARIO :
#  Au début de l'année scolaire, le secrétaire uploade
#  le fichier Excel GestionScolaire_SIGC.xlsx.
#  Ce use case lit les 6 feuilles dans l'ordre et insère
#  (ou met à jour) les données en base.
#
#  ORDRE D'IMPORT OBLIGATOIRE (respecter les clés étrangères) :
#  1. Specialites   (pas de dépendance)
#  2. Etudiants     (dépend de Specialites)
#  3. Paiements     (dépend de Etudiants et Specialites)
#  4. QR_Codes      (dépend de Etudiants et Paiements)
#  5. Notifications (dépend de Etudiants et Paiements)
#  6. Calendrier    (pas de dépendance)
#
#  COUCHE : Application
# ============================================================

import io                      # Pour lire les bytes du fichier uploadé
from decimal import Decimal, InvalidOperation
from typing import List

import openpyxl                # Pour lire les fichiers Excel

from app.Domain.entities import (
    SpecialiteDomaine, EtudiantDomaine, PaiementDomaine,
    QRCodeDomaine, NotificationDomaine,
)
from app.Domain.value_objects import (
    Matricule, Montant, Email, Telephone,
    Sexe, LienParent, StatutPaiement, StatutQRCode,
    TypeNotification, Niveau,
)
from app.Domain.interfaces import (
    ISpecialiteRepository,
    IEtudiantRepository,
    IPaiementRepository,
    IQRCodeRepository,
    INotificationRepository,
)
from app.Domain.exceptions import FichierExcelInvalideError
from app.Application.DTOs.schemas import (
    ImportExcelReponseDTO,
    ResultatFeuilleDTO,
)


# ── Fonctions de conversion ──────────────────────────────────
# Ces petites fonctions convertissent une valeur brute Excel
# en type Python propre, sans lever d'exception si la valeur est nulle.

def _str(valeur) -> str:
    """Convertit en string, retourne '' si None."""
    return str(valeur).strip() if valeur not in (None, "") else ""

def _int(valeur) -> int:
    """Convertit en entier, retourne 0 si impossible."""
    try:
        return int(valeur)
    except (TypeError, ValueError):
        return 0

def _decimal(valeur) -> Decimal:
    """Convertit en Decimal, retourne 0.00 si impossible."""
    try:
        return Decimal(str(valeur)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError):
        return Decimal("0.00")

def _bool_oui(valeur) -> bool:
    """Convertit 'OUI'/'NON' en True/False."""
    return str(valeur).strip().upper() == "OUI"

def _statut_paiement(valeur: str) -> StatutPaiement:
    """Convertit le texte Excel en enum StatutPaiement."""
    correspondances = {
        "PAYÉ":       StatutPaiement.PAYE,
        "PAYE":       StatutPaiement.PAYE,
        "EN ATTENTE": StatutPaiement.EN_ATTENTE,
        "EN RETARD":  StatutPaiement.EN_RETARD,
        "PARTIEL":    StatutPaiement.PARTIEL,
    }
    # .get(cle, valeur_par_defaut) → retourne EN_ATTENTE si non reconnu
    return correspondances.get(
        str(valeur).strip().upper(),
        StatutPaiement.EN_ATTENTE
    )

def _statut_qr(valeur: str) -> StatutQRCode:
    correspondances = {
        "ACTIF":    StatutQRCode.ACTIF,
        "EXPIRÉ":   StatutQRCode.EXPIRE,
        "SUSPENDU": StatutQRCode.SUSPENDU,
    }
    return correspondances.get(str(valeur).strip().upper(), StatutQRCode.ACTIF)

def _niveau(valeur) -> Niveau:
    """Convertit un entier 1-5 en enum Niveau."""
    try:
        return Niveau(int(valeur))
    except (ValueError, TypeError):
        return Niveau.UN  # Valeur par défaut si invalide

def _lien_parent(valeur: str) -> LienParent:
    v = str(valeur).strip().lower()
    if "mère" in v or "mere" in v:
        return LienParent.MERE
    if "père" in v or "pere" in v:
        return LienParent.PERE
    return LienParent.TUTEUR

def _sexe(valeur: str) -> Sexe:
    return Sexe.FEMININ if str(valeur).strip().upper() == "F" else Sexe.MASCULIN

def _type_notif(valeur: str) -> TypeNotification:
    v = str(valeur).strip().upper()
    if "CONFIRMATION" in v:
        return TypeNotification.CONFIRMATION
    if "PARTIEL" in v:
        return TypeNotification.PARTIEL
    if "RELANCE" in v:
        return TypeNotification.RELANCE
    return TypeNotification.RAPPEL


# ============================================================
#  CAS D'USAGE : ImporterExcelUseCase
# ============================================================
class ImporterExcelUseCase:
    """
    Importe le fichier Excel SIGC complet en base de données.
    Traite les 6 feuilles dans le bon ordre.
    """

    def __init__(
        self,
        specialite_repo: ISpecialiteRepository,
        etudiant_repo:   IEtudiantRepository,
        paiement_repo:   IPaiementRepository,
        qr_repo:         IQRCodeRepository,
        notif_repo:      INotificationRepository,
        calendrier_repo = None,
    ):
        self._specialite_repo = specialite_repo
        self._etudiant_repo   = etudiant_repo
        self._paiement_repo   = paiement_repo
        self._qr_repo         = qr_repo
        self._notif_repo      = notif_repo
        self._calendrier_repo = calendrier_repo

    async def executer(
        self,
        fichier_bytes: bytes,
        nom_fichier:   str,
    ) -> ImportExcelReponseDTO:
        """
        Lit le fichier Excel et importe chaque feuille.

        fichier_bytes = contenu binaire du fichier uploadé
        nom_fichier   = nom du fichier (pour valider l'extension)
        """

        # Valider l'extension du fichier
        if not nom_fichier.lower().endswith((".xlsx", ".xlsm")):
            raise FichierExcelInvalideError(nom_fichier)

        # Ouvrir le classeur Excel depuis les bytes en mémoire
        # io.BytesIO transforme des bytes en "fichier virtuel"
        # data_only=True → lire les VALEURS des cellules, pas les formules
        try:
            classeur = openpyxl.load_workbook(
                filename  = io.BytesIO(fichier_bytes),
                data_only = True,
            )
        except Exception as e:
            raise FichierExcelInvalideError(f"{nom_fichier} — {e}")

        # ── Traiter les feuilles dans l'ordre ───────────────
        # On appelle chaque méthode privée _importer_xxx()
        resultats: List[ResultatFeuilleDTO] = []

        resultats.append(await self._importer_specialites(classeur))
        resultats.append(await self._importer_etudiants(classeur))
        resultats.append(await self._importer_paiements(classeur))
        resultats.append(await self._importer_qr_codes(classeur))
        resultats.append(await self._importer_notifications(classeur))
        resultats.append(await self._importer_calendrier(classeur))

        # Compter le total et les erreurs
        total         = sum(r.inseres + r.mis_a_jour for r in resultats)
        total_erreurs = sum(r.erreurs for r in resultats)

        return ImportExcelReponseDTO(
            succes       = total_erreurs == 0,
            message      = (
                "Import terminé avec succès ✅"
                if total_erreurs == 0
                else f"Import terminé avec {total_erreurs} erreur(s) ⚠️"
            ),
            total_lignes = total,
            resultats    = resultats,
        )

    # ── Méthodes privées — une par feuille ──────────────────

    async def _importer_specialites(self, classeur) -> ResultatFeuilleDTO:
        """
        Lit la feuille 'Specialites' et insère/met à jour en base.
        Les lignes de données commencent à la ligne 3 (1=titre, 2=en-tête).
        """
        NOM_FEUILLE = "Specialites"
        resultat = ResultatFeuilleDTO(
            feuille=NOM_FEUILLE, inseres=0, mis_a_jour=0, erreurs=0
        )

        # Vérifier que la feuille existe dans le fichier
        if NOM_FEUILLE not in classeur.sheetnames:
            resultat.details.append(f"Feuille '{NOM_FEUILLE}' absente.")
            return resultat

        feuille = classeur[NOM_FEUILLE]

        # iter_rows(min_row=3) → on commence à la ligne 3 (saute titre + en-tête)
        # values_only=True → on reçoit des valeurs brutes, pas des objets Cell
        for ligne in feuille.iter_rows(min_row=3, values_only=True):

            # Si la ligne est vide ou n'a pas d'ID valide → on saute
            if not ligne[0] or not str(ligne[0]).startswith("SP-"):
                continue

            try:
                # Construire l'entité Domaine depuis les valeurs Excel
                # ligne[0] = colonne A, ligne[1] = colonne B, etc.
                specialite = SpecialiteDomaine(
                    id_specialite    = _str(ligne[0]),
                    code             = _str(ligne[1]),
                    nom_specialite   = _str(ligne[2]),
                    departement      = _str(ligne[3]),
                    niveau           = _niveau(ligne[4]),
                    duree_ans        = _int(ligne[5]),
                    annee_academique = _str(ligne[6]),
                    tranche_1        = Montant(_decimal(ligne[7])),
                    date_limite_t1   = _str(ligne[8]),
                    tranche_2        = Montant(_decimal(ligne[9])),
                    date_limite_t2   = _str(ligne[10]),
                    tranche_3        = Montant(_decimal(ligne[11])),
                    date_limite_t3   = _str(ligne[12]),
                )

                # sauvegarder() gère le upsert (insert ou update)
                await self._specialite_repo.sauvegarder(specialite)

                # On ne sait pas si c'était un insert ou update ici,
                # on compte dans inseres pour simplifier
                resultat.inseres += 1

            except Exception as erreur:
                # On ne bloque pas tout l'import pour une ligne en erreur.
                # On note l'erreur et on continue.
                resultat.erreurs += 1
                resultat.details.append(f"Ligne {ligne[0]} : {erreur}")

        return resultat

    async def _importer_etudiants(self, classeur) -> ResultatFeuilleDTO:
        """Lit la feuille 'Etudiants' et insère/met à jour en base."""
        NOM_FEUILLE = "Etudiants"
        resultat = ResultatFeuilleDTO(
            feuille=NOM_FEUILLE, inseres=0, mis_a_jour=0, erreurs=0
        )

        if NOM_FEUILLE not in classeur.sheetnames:
            resultat.details.append(f"Feuille '{NOM_FEUILLE}' absente.")
            return resultat

        feuille = classeur[NOM_FEUILLE]

        for ligne in feuille.iter_rows(min_row=3, values_only=True):
            if not ligne[0] or not str(ligne[0]).startswith("ETU-"):
                continue
            try:
                # Vérifier si l'étudiant existe déjà
                existant = await self._etudiant_repo.trouver_par_id(_str(ligne[0]))

                etudiant = EtudiantDomaine(
                    id_etudiant        = _str(ligne[0]),
                    matricule          = Matricule(_str(ligne[1])),
                    nom                = _str(ligne[2]),
                    prenom             = _str(ligne[3]),
                    date_naissance     = _str(ligne[4]) or None,
                    sexe               = _sexe(ligne[5]),
                    id_specialite      = _str(ligne[6]),
                    code_specialite    = _str(ligne[7]),
                    niveau             = _niveau(ligne[8]),
                    annee_academique   = _str(ligne[9]),
                    email_etudiant     = Email(_str(ligne[10])) if _str(ligne[10]) else None,
                    telephone_etudiant = Telephone(_str(ligne[11])) if _str(ligne[11]) else None,
                    nom_parent         = _str(ligne[12]),
                    prenom_parent      = _str(ligne[13]),
                    lien_parent        = _lien_parent(ligne[14]),
                    telephone_parent   = Telephone(_str(ligne[15])) if _str(ligne[15]) else None,
                    email_parent       = Email(_str(ligne[16])) if _str(ligne[16]) else None,
                )

                await self._etudiant_repo.sauvegarder(etudiant)
                if existant:
                    resultat.mis_a_jour += 1
                else:
                    resultat.inseres += 1

            except Exception as erreur:
                resultat.erreurs += 1
                resultat.details.append(f"Ligne {ligne[0]} : {erreur}")

        return resultat

    async def _importer_paiements(self, classeur) -> ResultatFeuilleDTO:
        """Lit la feuille 'Paiements' et insère/met à jour en base."""
        NOM_FEUILLE = "Paiements"
        resultat = ResultatFeuilleDTO(
            feuille=NOM_FEUILLE, inseres=0, mis_a_jour=0, erreurs=0
        )

        if NOM_FEUILLE not in classeur.sheetnames:
            resultat.details.append(f"Feuille '{NOM_FEUILLE}' absente.")
            return resultat

        feuille = classeur[NOM_FEUILLE]

        for ligne in feuille.iter_rows(min_row=3, values_only=True):
            if not ligne[0] or not str(ligne[0]).startswith("PAY-"):
                continue
            try:
                existant = await self._paiement_repo.trouver_par_id(_str(ligne[0]))
                m_attendu = Montant(_decimal(ligne[6]))
                m_paye    = Montant(_decimal(ligne[7]))

                paiement = PaiementDomaine(
                    id_paiement    = _str(ligne[0]),
                    id_etudiant    = _str(ligne[1]),
                    id_specialite  = _str(ligne[3]),
                    niveau         = _int(ligne[4]),
                    numero_tranche = _int(ligne[5]),
                    montant_attendu= m_attendu,
                    montant_paye   = m_paye,
                    date_paiement  = _str(ligne[8]) or None,
                    date_limite    = _str(ligne[9]),
                    statut         = _statut_paiement(ligne[10]),
                    qr_code_genere = _bool_oui(ligne[14]),
                    notif_envoyee  = _bool_oui(ligne[15]),
                    observations   = _str(ligne[16]) or None,
                )

                await self._paiement_repo.sauvegarder(paiement)
                if existant:
                    resultat.mis_a_jour += 1
                else:
                    resultat.inseres += 1

            except Exception as erreur:
                resultat.erreurs += 1
                resultat.details.append(f"Ligne {ligne[0]} : {erreur}")

        return resultat

    async def _importer_qr_codes(self, classeur) -> ResultatFeuilleDTO:
        """Lit la feuille 'QR_Codes' et insère/met à jour en base."""
        NOM_FEUILLE = "QR_Codes"
        resultat = ResultatFeuilleDTO(
            feuille=NOM_FEUILLE, inseres=0, mis_a_jour=0, erreurs=0
        )

        if NOM_FEUILLE not in classeur.sheetnames:
            resultat.details.append(f"Feuille '{NOM_FEUILLE}' absente.")
            return resultat

        feuille = classeur[NOM_FEUILLE]

        for ligne in feuille.iter_rows(min_row=3, values_only=True):
            if not ligne[0] or not str(ligne[0]).startswith("QR-"):
                continue
            try:
                qr = QRCodeDomaine(
                    id_qrcode       = _str(ligne[0]),
                    id_etudiant     = _str(ligne[1]),
                    id_paiement     = None,
                    id_specialite   = _str(ligne[3]),
                    niveau          = _int(ligne[4]),
                    date_generation = _str(ligne[5]),
                    valide_jusqua   = _str(ligne[6]),
                    statut          = _statut_qr(ligne[7]),
                )
                await self._qr_repo.sauvegarder(qr)
                resultat.inseres += 1

            except Exception as erreur:
                resultat.erreurs += 1
                resultat.details.append(f"Ligne {ligne[0]} : {erreur}")

        return resultat

    async def _importer_notifications(self, classeur) -> ResultatFeuilleDTO:
        """Lit la feuille 'Notifications' et insère/met à jour en base."""
        NOM_FEUILLE = "Notifications"
        resultat = ResultatFeuilleDTO(
            feuille=NOM_FEUILLE, inseres=0, mis_a_jour=0, erreurs=0
        )

        if NOM_FEUILLE not in classeur.sheetnames:
            resultat.details.append(f"Feuille '{NOM_FEUILLE}' absente.")
            return resultat

        feuille = classeur[NOM_FEUILLE]

        for ligne in feuille.iter_rows(min_row=3, values_only=True):
            if not ligne[0] or not str(ligne[0]).startswith("NOTIF-"):
                continue
            try:
                notif = NotificationDomaine(
                    id_notification   = _str(ligne[0]),
                    id_paiement       = _str(ligne[1]),
                    id_etudiant       = _str(ligne[2]),
                    nom_parent        = _str(ligne[4]),
                    contact_parent    = _str(ligne[5]),
                    type_notification = _type_notif(ligne[6]),
                    message           = _str(ligne[7]),
                    date_envoi        = _str(ligne[8]) or None,
                    canal             = _str(ligne[9]),
                    statut_envoi      = _str(ligne[10]),
                )
                await self._notif_repo.sauvegarder(notif)
                resultat.inseres += 1

            except Exception as erreur:
                resultat.erreurs += 1
                resultat.details.append(f"Ligne {ligne[0]} : {erreur}")

        return resultat

    async def _importer_calendrier(self, classeur) -> ResultatFeuilleDTO:
        """
        Lit la feuille 'Calendrier_Niveaux' et insère/met à jour en base.

        Colonnes attendues (à partir de la ligne 3) :
        A=niveau, B=groupe, C=demarrage_academique, D=tranche_1_limite,
        E=tranche_2_limite, F=tranche_3_limite, G=condition_demarrage,
        H=mois_debut, I=mois_fin
        """
        NOM_FEUILLE = "Calendrier_Niveaux"
        resultat = ResultatFeuilleDTO(
            feuille=NOM_FEUILLE, inseres=0, mis_a_jour=0, erreurs=0
        )

        if self._calendrier_repo is None:
            resultat.details.append(
                "Calendrier_repo non configuré — feuille ignorée."
            )
            return resultat

        if NOM_FEUILLE not in classeur.sheetnames:
            resultat.details.append(f"Feuille '{NOM_FEUILLE}' absente.")
            return resultat

        feuille = classeur[NOM_FEUILLE]

        for ligne in feuille.iter_rows(min_row=3, values_only=True):
            niveau = _int(ligne[0])
            if not niveau:
                continue
            try:
                existant = await self._calendrier_repo.trouver_par_niveau(niveau)

                await self._calendrier_repo.sauvegarder(
                    niveau                = niveau,
                    groupe                = _str(ligne[1]),
                    demarrage_academique  = _str(ligne[2]),
                    tranche_1_limite      = _str(ligne[3]),
                    tranche_2_limite      = _str(ligne[4]),
                    tranche_3_limite      = _str(ligne[5]),
                    condition_demarrage   = _str(ligne[6]),
                    mois_debut            = _str(ligne[7]),
                    mois_fin              = _str(ligne[8]),
                )

                if existant:
                    resultat.mis_a_jour += 1
                else:
                    resultat.inseres += 1

            except Exception as erreur:
                resultat.erreurs += 1
                resultat.details.append(f"Ligne niveau={ligne[0]} : {erreur}")

        return resultat
