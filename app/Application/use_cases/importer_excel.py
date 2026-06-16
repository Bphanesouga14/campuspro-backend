

#  RÔLE : Orchestrer l'import du fichier Excel en base.
#
#  ORDRE D'IMPORT OBLIGATOIRE (contraintes de clés étrangères) :
#  1. Spécialités  → doivent exister avant les étudiants
#  2. Étudiants    → doivent exister avant les paiements
#  3. Paiements    → doivent exister avant QR codes et notifs
#  4. QR Codes     → dépendent des étudiants et paiements
#  5. Notifications→ dépendent des étudiants et paiements
#  6. Calendrier   → indépendant, peut aller n'importe quand
#
#  COUCHE : Application
# ============================================================

import io
from decimal import Decimal, InvalidOperation
from typing  import List, Tuple

from openpyxl import load_workbook

from app.Domain.entities      import SpecialiteDomaine, EtudiantDomaine, PaiementDomaine
from app.Domain.value_objects import (
    Montant, Matricule, Telephone, Email,
    Sexe, LienParent, StatutPaiement, Niveau
)
from app.Domain.interfaces    import (
    ISpecialiteRepository,
    IEtudiantRepository,
    IPaiementRepository,
)
from app.Domain.exceptions    import FichierExcelInvalideError
from app.Application.DTOs.schemas import ImportExcelOut, ResultatFeuille


# ── FONCTIONS UTILITAIRES ────────────────────────────────────
# Ces fonctions convertissent les valeurs brutes d'Excel
# en types Python propres.

def _texte(valeur) -> str:
    """Convertit n'importe quelle valeur en texte propre (sans espaces)."""
    return str(valeur).strip() if valeur not in (None, "", "None") else ""

def _entier(valeur) -> int:
    """Convertit en entier, retourne 0 si impossible."""
    try:
        return int(valeur)
    except (TypeError, ValueError):
        return 0

def _decimal(valeur) -> Decimal:
    """Convertit en Decimal financier, retourne 0.00 si impossible."""
    try:
        return Decimal(str(valeur)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError):
        return Decimal("0.00")

def _booleen(valeur) -> bool:
    """Convertit "OUI"/"NON" en True/False."""
    return str(valeur).strip().upper() == "OUI"

def _statut_paiement(valeur: str) -> StatutPaiement:
    """Convertit le texte du statut en Enum StatutPaiement."""
    table = {
        "PAYÉ":       StatutPaiement.PAYE,
        "PAYE":       StatutPaiement.PAYE,
        "EN ATTENTE": StatutPaiement.EN_ATTENTE,
        "EN RETARD":  StatutPaiement.EN_RETARD,
        "PARTIEL":    StatutPaiement.PARTIEL,
    }
    return table.get(str(valeur).strip().upper(), StatutPaiement.EN_ATTENTE)

def _sexe(valeur: str) -> Sexe:
    return Sexe.FEMININ if str(valeur).strip().upper() == "F" else Sexe.MASCULIN

def _lien(valeur: str) -> LienParent:
    v = str(valeur).strip().lower()
    if "mère" in v or "mere" in v: return LienParent.MERE
    if "père" in v or "pere" in v: return LienParent.PERE
    return LienParent.TUTEUR

def _niveau(valeur) -> Niveau:
    try:
        return Niveau(_entier(valeur))
    except ValueError:
        return Niveau.UN   # Valeur par défaut si niveau inconnu


# ════════════════════════════════════════════════════════════
#  CAS D'USAGE PRINCIPAL
# ════════════════════════════════════════════════════════════

class ImporterExcelUseCase:
    """
    Cas d'usage : Importer le fichier Excel GestionScolaire_SIGC.xlsx.

    Lit chaque feuille, construit les entités domaine,
    et sauvegarde en base via les repositories.
    Retourne un rapport détaillé de l'import.
    """

    def __init__(
        self,
        specialite_repo: ISpecialiteRepository,
        etudiant_repo:   IEtudiantRepository,
        paiement_repo:   IPaiementRepository,
    ):
        self._specialite_repo = specialite_repo
        self._etudiant_repo   = etudiant_repo
        self._paiement_repo   = paiement_repo

    async def executer(
        self,
        contenu_fichier: bytes,
        nom_fichier:     str,
    ) -> ImportExcelOut:
        """
        Lance l'import complet du fichier Excel.

        contenu_fichier = le fichier en mémoire (bytes)
        nom_fichier     = le nom du fichier (pour les messages d'erreur)
        """

        # ── Vérification du format du fichier ─────────────────
        if not nom_fichier.endswith((".xlsx", ".xlsm")):
            raise FichierExcelInvalideError(nom_fichier)

        # ── Charger le classeur Excel en mémoire ──────────────
        # io.BytesIO transforme les bytes en "fichier virtuel"
        # que openpyxl peut lire sans écrire sur le disque
        try:
            classeur = load_workbook(
                filename   = io.BytesIO(contenu_fichier),
                data_only  = True   # data_only=True → lire les valeurs calculées
                                    # et non les formules Excel brutes
            )
        except Exception as e:
            raise FichierExcelInvalideError(f"{nom_fichier} : {e}")

        # ── Importer dans l'ordre des dépendances ─────────────
        resultats = []
        total     = 0

        # 1. Spécialités en premier (les étudiants y font référence)
        res = await self._importer_specialites(classeur)
        resultats.append(res)
        total += res.inseres + res.mis_a_jour

        # 2. Étudiants (les paiements y font référence)
        res = await self._importer_etudiants(classeur)
        resultats.append(res)
        total += res.inseres + res.mis_a_jour

        # 3. Paiements (QR codes et notifications y font référence)
        res = await self._importer_paiements(classeur)
        resultats.append(res)
        total += res.inseres + res.mis_a_jour

        # Compter les erreurs totales
        erreurs_totales = sum(r.erreurs for r in resultats)

        return ImportExcelOut(
            succes       = erreurs_totales == 0,
            message      = (
                f"Import terminé. {total} lignes traitées."
                if erreurs_totales == 0
                else f"Import terminé avec {erreurs_totales} erreur(s). "
                     f"Vérifiez les détails ci-dessous."
            ),
            total_lignes = total,
            resultats    = resultats,
        )

    # ── IMPORTEURS PAR FEUILLE ───────────────────────────────

    async def _importer_specialites(self, classeur) -> ResultatFeuille:
        """
        Lit la feuille 'Specialites' et sauvegarde en base.
        Retourne un rapport : insérés / mis à jour / erreurs.
        """
        resultat = ResultatFeuille(
            feuille="Specialites", inseres=0, mis_a_jour=0, erreurs=0
        )

        # Vérifier que la feuille existe dans le fichier
        if "Specialites" not in classeur.sheetnames:
            resultat.details.append("Feuille 'Specialites' absente du fichier.")
            resultat.erreurs += 1
            return resultat

        feuille = classeur["Specialites"]

        # Lire à partir de la ligne 3 (les 2 premières sont l'en-tête)
        for ligne in feuille.iter_rows(min_row=3, values_only=True):

            # Ignorer les lignes vides ou qui ne commencent pas par "SP-"
            if not ligne[0] or not str(ligne[0]).startswith("SP-"):
                continue

            try:
                # Lire les valeurs de la ligne Excel
                id_sp = _texte(ligne[0])   # Col A : ID Spécialité
                # Colonne :   A      B      C       D        E       F        G
                # Index   :   0      1      2       3        4       5        6
                #              id    code   nom     dept     niveau  durée    année
                #           H(7)   I(8)   J(9)   K(10)  L(11)  M(12)

                # Construire le montant total = T1 + T2 + T3
                t1 = Montant(_decimal(ligne[7]))
                t2 = Montant(_decimal(ligne[9]))
                t3 = Montant(_decimal(ligne[11]))

                # Construire l'entité domaine SpecialiteDomaine
                specialite = SpecialiteDomaine(
                    id_specialite    = id_sp,
                    code             = _texte(ligne[1]),
                    nom_specialite   = _texte(ligne[2]),
                    departement      = _texte(ligne[3]),
                    niveau           = _niveau(ligne[4]),
                    duree_ans        = _entier(ligne[5]) or 1,
                    annee_academique = _texte(ligne[6]),
                    tranche_1        = t1,
                    date_limite_t1   = _texte(ligne[8]),
                    tranche_2        = t2,
                    date_limite_t2   = _texte(ligne[10]),
                    tranche_3        = t3,
                    date_limite_t3   = _texte(ligne[12]),
                )

                # Vérifier si elle existe déjà en base (upsert)
                existante = await self._specialite_repo.trouver_par_id(id_sp)
                await self._specialite_repo.sauvegarder(specialite)

                # Compter : nouveau ou mise à jour ?
                if existante:
                    resultat.mis_a_jour += 1
                else:
                    resultat.inseres += 1

            except Exception as e:
                # Une erreur sur une ligne n'arrête pas tout l'import
                resultat.erreurs += 1
                resultat.details.append(f"Spécialité {ligne[0]} : {str(e)}")

        return resultat

    async def _importer_etudiants(self, classeur) -> ResultatFeuille:
        """Lit la feuille 'Etudiants' et sauvegarde en base."""
        resultat = ResultatFeuille(
            feuille="Etudiants", inseres=0, mis_a_jour=0, erreurs=0
        )

        if "Etudiants" not in classeur.sheetnames:
            resultat.details.append("Feuille 'Etudiants' absente du fichier.")
            resultat.erreurs += 1
            return resultat

        feuille = classeur["Etudiants"]

        for ligne in feuille.iter_rows(min_row=3, values_only=True):
            if not ligne[0] or not str(ligne[0]).startswith("ETU-"):
                continue

            try:
                # Colonnes de la feuille Etudiants :
                # A(0)=id  B(1)=matricule  C(2)=nom  D(3)=prenom
                # E(4)=ddn  F(5)=sexe  G(6)=id_sp  H(7)=code_sp
                # I(8)=niveau  J(9)=annee  K(10)=email_et  L(11)=tel_et
                # M(12)=nom_parent  N(13)=prenom_parent  O(14)=lien
                # P(15)=tel_parent  Q(16)=email_parent

                matricule_str = _texte(ligne[1])
                # Valider le matricule via le Value Object
                # Si le format est invalide → exception capturée plus bas
                matricule_vo = Matricule(matricule_str)

                # Construire les Value Objects optionnels
                email_et = Email(_texte(ligne[10]))   if _texte(ligne[10])  else None
                tel_et   = Telephone(_texte(ligne[11])) if _texte(ligne[11]) else None
                tel_par  = Telephone(_texte(ligne[15])) if _texte(ligne[15]) else None
                email_par= Email(_texte(ligne[16]))   if _texte(ligne[16])  else None

                etudiant = EtudiantDomaine(
                    id_etudiant        = _texte(ligne[0]),
                    matricule          = matricule_vo,
                    nom                = _texte(ligne[2]).upper(),
                    prenom             = _texte(ligne[3]).capitalize(),
                    date_naissance     = _texte(ligne[4]) or None,
                    sexe               = _sexe(ligne[5]),
                    id_specialite      = _texte(ligne[6]),
                    code_specialite    = _texte(ligne[7]),
                    niveau             = _niveau(ligne[8]),
                    annee_academique   = _texte(ligne[9]),
                    email_etudiant     = email_et,
                    telephone_etudiant = tel_et,
                    nom_parent         = _texte(ligne[12]).upper(),
                    prenom_parent      = _texte(ligne[13]).capitalize(),
                    lien_parent        = _lien(ligne[14]),
                    telephone_parent   = tel_par,
                    email_parent       = email_par,
                )

                existant = await self._etudiant_repo.trouver_par_id(etudiant.id_etudiant)
                await self._etudiant_repo.sauvegarder(etudiant)

                if existant:
                    resultat.mis_a_jour += 1
                else:
                    resultat.inseres += 1

            except Exception as e:
                resultat.erreurs += 1
                resultat.details.append(f"Étudiant {ligne[0]} : {str(e)}")

        return resultat

    async def _importer_paiements(self, classeur) -> ResultatFeuille:
        """Lit la feuille 'Paiements' et sauvegarde en base."""
        resultat = ResultatFeuille(
            feuille="Paiements", inseres=0, mis_a_jour=0, erreurs=0
        )

        if "Paiements" not in classeur.sheetnames:
            resultat.details.append("Feuille 'Paiements' absente du fichier.")
            resultat.erreurs += 1
            return resultat

        feuille = classeur["Paiements"]

        for ligne in feuille.iter_rows(min_row=3, values_only=True):
            if not ligne[0] or not str(ligne[0]).startswith("PAY-"):
                continue

            try:
                # Colonnes de la feuille Paiements :
                # A(0)=id_pay  B(1)=id_et  C(2)=nom_complet  D(3)=id_sp
                # E(4)=niveau  F(5)=tranche  G(6)=attendu  H(7)=paye
                # I(8)=date_paiement  J(9)=date_limite  K(10)=statut
                # L(11)=reste  M(12)=cumul  N(13)=reste_global
                # O(14)=qr_genere  P(15)=notif_envoyee  Q(16)=observations

                paiement = PaiementDomaine(
                    id_paiement     = _texte(ligne[0]),
                    id_etudiant     = _texte(ligne[1]),
                    id_specialite   = _texte(ligne[3]),
                    niveau          = _entier(ligne[4]),
                    numero_tranche  = _entier(ligne[5]),
                    montant_attendu = Montant(_decimal(ligne[6])),
                    montant_paye    = Montant(_decimal(ligne[7])),
                    date_paiement   = _texte(ligne[8])  or None,
                    date_limite     = _texte(ligne[9]),
                    statut          = _statut_paiement(ligne[10]),
                    qr_code_genere  = _booleen(ligne[14]),
                    notif_envoyee   = _booleen(ligne[15]),
                    observations    = _texte(ligne[16]) or None,
                )

                existant = await self._paiement_repo.trouver_par_id(paiement.id_paiement)
                await self._paiement_repo.sauvegarder(paiement)

                if existant:
                    resultat.mis_a_jour += 1
                else:
                    resultat.inseres += 1

            except Exception as e:
                resultat.erreurs += 1
                resultat.details.append(f"Paiement {ligne[0]} : {str(e)}")

        return resultat
